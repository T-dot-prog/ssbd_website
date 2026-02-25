"""
Helper function for SSbd website
"""
import numpy as np
from PIL import Image
import pandas as pd
import json
import openpyxl
from config import config
import streamlit as st

from asposeslidescloud import SlidesApi, Configuration

configuration = Configuration()
configuration.app_sid = "4658f1d5-dc13-452c-8229-d69d3f44f9aa"
configuration.app_key = "ab6ea99b2ce188cebc325677e161a650"
configuration.debug = False

slides_api = SlidesApi(configuration)
 
class HelperClass:
    def __init__(self, logo_data_path: str = config.LOGO_PATH, xml_file_path: str = config.XML_PATH):
        self.logo_path = logo_data_path
        self.xml_path = xml_file_path

    def load_logo(self) -> list:
        """Load Logo from Source link"""
        img = Image.open(fp=self.logo_path)
        img_to_array = np.array(img)
        return img_to_array
    
    def load_xlx(self) -> pd.DataFrame:
        """Load Xlxs from Source Link"""
        df = pd.read_excel(self.xml_path)
        return df.to_json()
    
    def extract_hyperlinks(self, column_name: str = "Formal Photo") -> dict:
        """
        Extract hyperlink URLs from a hyperlinked column in the Excel file
        using openpyxl. Returns a dict keyed by string row index (matches pandas).
        """
        wb = openpyxl.load_workbook(self.xml_path)
        ws = wb.worksheets[0]

        # Find the column index by header name in row 1
        header_col_idx = None
        for cell in ws[1]:
            if cell.value == column_name:
                header_col_idx = cell.column
                break

        if header_col_idx is None:
            return {}

        hyperlink_map = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
            cell = row[header_col_idx - 1]
            if cell.hyperlink and cell.hyperlink.target:
                hyperlink_map[str(row_idx)] = cell.hyperlink.target
            else:
                # Fallback to cell text value if no hyperlink
                hyperlink_map[str(row_idx)] = cell.value

        return hyperlink_map

    def convert_into_dict(self) -> dict[str, str]:
        """Convert dataframe into dictionary, with real hyperlinks for Formal Photo"""
        data = self.load_xlx()
        data_dict = json.loads(data)

        # Overwrite Formal Photo values with actual hyperlink URLs
        try:
            hyperlink_map = self.extract_hyperlinks(column_name="Formal Photo")
            if hyperlink_map:
                data_dict["Formal Photo"] = hyperlink_map
        except Exception as e:
            pass  # Leave Formal Photo as-is if extraction fails

        return data_dict
    
    def name_to_pdf(self, _id: str) -> bytes:
        """Names to pdf bytes"""
        with open(f"pdf/{_id}.pdf", "rb") as f:
            data = f.read()
        return data

    def drivelink_to_image(self, image_id: str) -> bytes:
        """Function to get drivelink to image"""
        import requests
        url = f"https://drive.google.com/uc?export=download&id={image_id}"
        response = requests.get(url)
        return response.content

    @staticmethod
    def convert_drive_url(url: str) -> str:
        """Convert a Google Drive shareable/view link to a direct download link."""
        if not url or not isinstance(url, str):
            return url
        if "drive.google.com/file/d/" in url:
            file_id = url.split("/file/d/")[1].split("/")[0]
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        if "drive.google.com/open?id=" in url:
            file_id = url.split("open?id=")[1].split("&")[0]
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        return url  # Already a direct link or unknown format


# Theme Management Functions
def initialize_theme():
    """Initialize theme in session state if not already present"""
    if "theme" not in st.session_state:
        st.session_state.theme = "dark"


def toggle_theme():
    """Toggle between light and dark themes"""
    st.session_state.theme = "dark" if st.session_state.theme == "light" else "light"


def get_current_theme():
    """Get current theme ('light' or 'dark')"""
    return st.session_state.get("theme", "dark")


def is_dark_theme():
    """Check if current theme is dark"""
    return get_current_theme() == "dark"


def render_theme_toggle():
    """Render theme toggle button in fixed position"""
    st.markdown("""
    <style>
        .theme-toggle-container {
            position: fixed;
            top: 15px;
            right: 25px;
            z-index: 9999;
        }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([10, 1, 0.2])
    with col3:
        if st.button("🌙" if is_dark_theme() else "☀️", key="theme_switch"):
            toggle_theme()
            st.rerun()


def get_dark_theme_css():
    """Return dark theme CSS"""
    return """
    <style>
        .stApp { 
            background-color: #0a0e27; 
            color: white; 
            font-family: 'Times New Roman', Times, serif !important; 
        }
        
        .main {
            background: linear-gradient(135deg, #0a0e27 0%, #1a1f3a 25%, #2d1b4e 50%, #1a1f3a 75%, #0a0e27 100%);
            min-height: 100vh;
            padding: 0;
            position: relative;
            overflow: hidden;
        }
        
        .main::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: radial-gradient(circle at 20% 50%, rgba(100, 181, 246, 0.1) 0%, transparent 50%),
                        radial-gradient(circle at 80% 80%, rgba(129, 199, 132, 0.1) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }
        
        .block-container {
            padding-top: 0;
            padding-bottom: 2rem;
            max-width: 100%;
            color: white;
            position: relative;
            z-index: 1;
        }
        
        h1, h2, h3, p, li, strong { 
            color: #f5f5f5 !important; 
            font-family: 'Times New Roman', Times, serif !important; 
        }
        
        .brand-title {
            font-family: 'Times New Roman', Times, serif;
            font-weight: 400;
            font-size: 3rem;
            color: #64b5f6;
            letter-spacing: 1px;
            margin: 0;
            text-transform: uppercase;
            text-shadow: 0 0 20px rgba(100, 181, 246, 0.3);
        }
        
        .header-container {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.15) 0%, rgba(255, 255, 255, 0.08) 100%);
            backdrop-filter: blur(30px);
            -webkit-backdrop-filter: blur(30px);
            border: 1px solid rgba(255, 255, 255, 0.25);
            border-radius: 30px;
            padding: 2.5rem 3rem;
            box-shadow: 
                0 8px 32px rgba(0, 0, 0, 0.3),
                inset 0 1px 1px rgba(255, 255, 255, 0.3),
                inset 0 -1px 1px rgba(0, 0, 0, 0.2);
            margin: 1.5rem;
            position: relative;
            overflow: hidden;
        }
        
        .input-card {
            background: linear-gradient(135deg, rgba(100, 181, 246, 0.15) 0%, rgba(129, 199, 132, 0.1) 100%);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 2px solid rgba(100, 181, 246, 0.3);
            border-radius: 25px;
            padding: 2.5rem;
            margin: 2rem 1.5rem;
            box-shadow: 
                0 10px 40px rgba(100, 181, 246, 0.2),
                inset 0 1px 2px rgba(255, 255, 255, 0.2);
            position: relative;
            overflow: hidden;
            transition: all 0.3s ease;
        }
        
        .input-card:hover {
            transform: translateY(-2px);
            box-shadow: 
                0 15px 50px rgba(100, 181, 246, 0.3),
                inset 0 1px 2px rgba(255, 255, 255, 0.3);
        }
        
        .input-label {
            font-family: 'Times New Roman', Times, serif;
            font-size: 1.1rem;
            font-weight: 600;
            color: #64b5f6;
            margin-bottom: 1rem;
            text-transform: uppercase;
            letter-spacing: 1px;
            text-shadow: 0 2px 10px rgba(100, 181, 246, 0.3);
        }
        
        .stTextInput > div > div > input {
            background: rgba(255, 255, 255, 0.1) !important;
            border: 2px solid rgba(100, 181, 246, 0.4) !important;
            border-radius: 15px !important;
            color: white !important;
            font-size: 1.1rem !important;
            padding: 1rem 1.5rem !important;
            font-family: 'Times New Roman', Times, serif !important;
            transition: all 0.3s ease !important;
            box-shadow: inset 0 2px 10px rgba(0, 0, 0, 0.2) !important;
        }
        
        .stTextInput > div > div > input:focus {
            background: rgba(255, 255, 255, 0.15) !important;
            border: 2px solid rgba(100, 181, 246, 0.8) !important;
            box-shadow: 
                inset 0 2px 10px rgba(0, 0, 0, 0.2),
                0 0 20px rgba(100, 181, 246, 0.4) !important;
            outline: none !important;
        }
        
        .stTextInput > div > div > input::placeholder {
            color: rgba(255, 255, 255, 0.5) !important;
        }
        
        /* Buttons (regular and download) */
        .stButton > button,
        .stDownloadButton > button,
        .stDownloadButton > a,
        .stDownloadButton a {
            background: linear-gradient(135deg, #64b5f6 0%, #81c784 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 15px !important;
            padding: 1rem 3rem !important;
            font-size: 1.1rem !important;
            font-weight: 600 !important;
            font-family: 'Times New Roman', Times, serif !important;
            text-transform: uppercase !important;
            letter-spacing: 1px !important;
            cursor: pointer !important;
            transition: all 0.3s ease !important;
            box-shadow: 
                0 8px 25px rgba(100, 181, 246, 0.4),
                inset 0 1px 2px rgba(255, 255, 255, 0.3) !important;
            position: relative !important;
            overflow: hidden !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            text-decoration: none !important;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        .stDownloadButton > a:hover,
        .stDownloadButton a:hover {
            transform: translateY(-3px) !important;
            box-shadow: 
                0 12px 35px rgba(100, 181, 246, 0.5),
                inset 0 1px 2px rgba(255, 255, 255, 0.4) !important;
        }
        
        .profile-card {
            background: linear-gradient(135deg, rgba(100, 181, 246, 0.15) 0%, rgba(129, 199, 132, 0.1) 100%);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 2px solid rgba(100, 181, 246, 0.3);
            border-radius: 24px;
            padding: 1.5rem;
            margin: 1rem 0 2rem 0;
            box-shadow: 0 10px 40px rgba(100, 181, 246, 0.2), inset 0 1px 2px rgba(255, 255, 255, 0.2);
        }
        
        .label {
            font-family: 'Times New Roman', Times, serif;
            color: #9ecbff;
            font-weight: 600;
            letter-spacing: 0.5px;
            text-transform: uppercase;
            font-size: 0.85rem;
        }
        
        .value {
            font-family: 'Times New Roman', Times, serif;
            color: #ffffff;
            font-weight: 600;
            font-size: 1.05rem;
        }
    </style>
    """


def get_light_theme_css():
    """Return light theme CSS"""
    return """
    <style>
        .stApp { 
            background-color: #f5f5f5; 
            color: #333; 
            font-family: 'Times New Roman', Times, serif !important; 
        }
        
        .main {
            background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 25%, #e8f4f8 50%, #f8f9fa 75%, #ffffff 100%);
            min-height: 100vh;
            padding: 0;
            position: relative;
            overflow: hidden;
        }
        
        .main::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: radial-gradient(circle at 20% 50%, rgba(100, 181, 246, 0.05) 0%, transparent 50%),
                        radial-gradient(circle at 80% 80%, rgba(129, 199, 132, 0.05) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }
        
        .block-container {
            padding-top: 0;
            padding-bottom: 2rem;
            max-width: 100%;
            color: #333;
            position: relative;
            z-index: 1;
        }
        
        h1, h2, h3, p, li, strong { 
            color: #1a1a1a !important; 
            font-family: 'Times New Roman', Times, serif !important; 
        }
        
        .brand-title {
            font-family: 'Times New Roman', Times, serif;
            font-weight: 400;
            font-size: 3rem;
            color: #2196F3;
            letter-spacing: 1px;
            margin: 0;
            text-transform: uppercase;
            text-shadow: 0 0 10px rgba(33, 150, 243, 0.1);
        }
        
        .header-container {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.8) 0%, rgba(248, 249, 250, 0.9) 100%);
            backdrop-filter: blur(30px);
            -webkit-backdrop-filter: blur(30px);
            border: 1px solid rgba(33, 150, 243, 0.2);
            border-radius: 30px;
            padding: 2.5rem 3rem;
            box-shadow: 
                0 8px 32px rgba(0, 0, 0, 0.1),
                inset 0 1px 1px rgba(255, 255, 255, 0.5);
            margin: 1.5rem;
            position: relative;
            overflow: hidden;
        }
        
        .input-card {
            background: linear-gradient(135deg, rgba(33, 150, 243, 0.08) 0%, rgba(76, 175, 80, 0.08) 100%);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 2px solid rgba(33, 150, 243, 0.2);
            border-radius: 25px;
            padding: 2.5rem;
            margin: 2rem 1.5rem;
            box-shadow: 
                0 10px 40px rgba(33, 150, 243, 0.1),
                inset 0 1px 2px rgba(255, 255, 255, 0.3);
            position: relative;
            overflow: hidden;
            transition: all 0.3s ease;
        }
        
        .input-card:hover {
            transform: translateY(-2px);
            box-shadow: 
                0 15px 50px rgba(33, 150, 243, 0.15),
                inset 0 1px 2px rgba(255, 255, 255, 0.4);
        }
        
        .input-label {
            font-family: 'Times New Roman', Times, serif;
            font-size: 1.1rem;
            font-weight: 600;
            color: #2196F3;
            margin-bottom: 1rem;
            text-transform: uppercase;
            letter-spacing: 1px;
            text-shadow: 0 1px 3px rgba(33, 150, 243, 0.1);
        }
        
        .stTextInput > div > div > input {
            background: rgba(255, 255, 255, 0.6) !important;
            border: 2px solid rgba(33, 150, 243, 0.3) !important;
            border-radius: 15px !important;
            color: #1a1a1a !important;
            font-size: 1.1rem !important;
            padding: 1rem 1.5rem !important;
            font-family: 'Times New Roman', Times, serif !important;
            transition: all 0.3s ease !important;
            box-shadow: inset 0 2px 10px rgba(0, 0, 0, 0.05) !important;
        }
        
        .stTextInput > div > div > input:focus {
            background: rgba(255, 255, 255, 0.9) !important;
            border: 2px solid rgba(33, 150, 243, 0.6) !important;
            box-shadow: 
                inset 0 2px 10px rgba(0, 0, 0, 0.05),
                0 0 20px rgba(33, 150, 243, 0.2) !important;
            outline: none !important;
        }
        
        .stTextInput > div > div > input::placeholder {
            color: rgba(26, 26, 26, 0.4) !important;
        }
        
        /* Buttons (regular and download) */
        .stButton > button,
        .stDownloadButton > button,
        .stDownloadButton > a,
        .stDownloadButton a {
            background: linear-gradient(135deg, #2196F3 0%, #4CAF50 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 15px !important;
            padding: 1rem 3rem !important;
            font-size: 1.1rem !important;
            font-weight: 600 !important;
            font-family: 'Times New Roman', Times, serif !important;
            text-transform: uppercase !important;
            letter-spacing: 1px !important;
            cursor: pointer !important;
            transition: all 0.3s ease !important;
            box-shadow: 
                0 8px 25px rgba(33, 150, 243, 0.2),
                inset 0 1px 2px rgba(255, 255, 255, 0.3) !important;
            position: relative !important;
            overflow: hidden !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            text-decoration: none !important;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        .stDownloadButton > a:hover,
        .stDownloadButton a:hover {
            transform: translateY(-3px) !important;
            box-shadow: 
                0 12px 35px rgba(33, 150, 243, 0.3),
                inset 0 1px 2px rgba(255, 255, 255, 0.4) !important;
        }
        
        .profile-card {
            background: linear-gradient(135deg, rgba(33, 150, 243, 0.08) 0%, rgba(76, 175, 80, 0.08) 100%);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 2px solid rgba(33, 150, 243, 0.2);
            border-radius: 24px;
            padding: 1.5rem;
            margin: 1rem 0 2rem 0;
            box-shadow: 0 10px 40px rgba(33, 150, 243, 0.1), inset 0 1px 2px rgba(255, 255, 255, 0.2);
        }
        
        .label {
            font-family: 'Times New Roman', Times, serif;
            color: #0D47A1 !important;
            font-weight: 700 !important;
            letter-spacing: 0.5px;
            text-transform: uppercase;
            font-size: 0.85rem;
        }
        
        .value {
            font-family: 'Times New Roman', Times, serif;
            color: #0D1B47 !important;
            font-weight: 700 !important;
            font-size: 1.05rem;
        }
    </style>
    """


helper = HelperClass()